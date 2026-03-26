"""
report_generator.py
===================
Professional Excel Report Builder.

Sheets produced
---------------
1. Executive Summary   – KPI scorecards + bar chart
2. Detailed Analysis   – full group-by table + pie chart
3. Monthly Trend       – time-series table + line chart  (if available)
4. Category Breakdown  – pivot cross-tab                 (if available)
5. Raw Data            – cleaned source data
"""

import io
import logging
from datetime import datetime
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import pandas as pd
import seaborn as sns
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Colour palette ─────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
ACCENT     = "E8612C"
LIGHT_GRAY = "F2F2F2"
WHITE      = "FFFFFF"
DARK_GRAY  = "404040"

_THIN   = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ── Style helpers ──────────────────────────────────────────────────────────────

def _hdr(ws, row, col, value, bg=DARK_BLUE, fg=WHITE, bold=True,
         size=11, h_align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=True)
    cell.border    = _BORDER
    return cell


def _val(ws, row, col, value, bold=False, size=10, h_align="left",
         bg=None, num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Arial", bold=bold, color=DARK_GRAY, size=size)
    cell.alignment = Alignment(horizontal=h_align, vertical="center")
    cell.border    = _BORDER
    if bg:      cell.fill          = PatternFill("solid", fgColor=bg)
    if num_fmt: cell.number_format = num_fmt
    return cell


def _auto_width(ws, min_w=10, max_w=40, padding=4):
    for col_cells in ws.columns:
        length = max(
            len(str(c.value)) if c.value is not None else 0
            for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = \
            min(max(length + padding, min_w), max_w)


# ══════════════════════════════════════════════════════════════════════════════
# REPORT GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

class ReportGenerator:
    """Build a multi-sheet formatted Excel report from analysis results.

    Parameters
    ----------
    analysis_results : dict returned by DataAnalyzer.analyze()
    raw_df           : cleaned DataFrame for the Raw Data sheet
    report_title     : title shown in the Executive Summary header
    group_col        : primary grouping column name
    value_col        : numeric KPI column name
    output_dir       : folder where the .xlsx will be saved
    """

    def __init__(self, analysis_results: dict, raw_df: pd.DataFrame,
                 report_title: str = "Business Performance Report",
                 group_col: str = "Group", value_col: str = "Value",
                 output_dir: str = "output"):
        self.results    = analysis_results
        self.raw_df     = raw_df
        self.title      = report_title
        self.group_col  = group_col
        self.value_col  = value_col
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    def generate(self) -> Path:
        """Build workbook, save it, return the file path."""
        logger.info("Generating Excel report …")
        self._build_executive_summary()
        self._build_detailed_analysis()
        if "monthly_trend" in self.results:
            self._build_monthly_trend()
        if "category_breakdown" in self.results:
            self._build_category_breakdown()
        self._build_raw_data()

        out = self.output_dir / f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.wb.save(out)
        logger.info("Report saved → %s", out)
        return out

    # ── Sheet 1 ───────────────────────────────────────────────────────────────
    def _build_executive_summary(self):
        ws = self.wb.create_sheet("Executive Summary")
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 40

        ws.merge_cells("A1:F1")
        c = ws["A1"]
        c.value     = f"📊  {self.title}"
        c.font      = Font(name="Arial", bold=True, size=18, color=WHITE)
        c.fill      = PatternFill("solid", fgColor=DARK_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A2:F2")
        s = ws["A2"]
        s.value     = f"Generated on {datetime.now().strftime('%d %B %Y, %H:%M')}"
        s.font      = Font(name="Arial", size=9, color=MID_BLUE)
        s.alignment = Alignment(horizontal="center")

        kpis = self.results.get("overall_kpis", {})
        row = 4
        _hdr(ws, row, 1, "Key Performance Indicators",
             bg=MID_BLUE, fg=WHITE, size=12, h_align="left")
        ws.merge_cells(f"A{row}:F{row}")
        row += 1

        for i, (label, value) in enumerate(kpis.items()):
            col = (i % 3) * 2 + 1
            if i > 0 and i % 3 == 0:
                row += 2
            for r, v, fgc, bgc, fsz in [
                (row,   label, WHITE,    MID_BLUE,   9),
                (row+1, value, DARK_BLUE, LIGHT_BLUE, 14),
            ]:
                cell = ws.cell(row=r, column=col, value=v)
                cell.font      = Font(name="Arial", bold=True, color=fgc, size=fsz)
                cell.fill      = PatternFill("solid", fgColor=bgc)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = _BORDER
                ws.row_dimensions[r+1].height = 30
                if col + 1 <= 6:
                    ws.merge_cells(start_row=r, start_column=col,
                                   end_row=r, end_column=col+1)

        top_df = self.results.get("top_performers", pd.DataFrame())
        if not top_df.empty:
            img = XLImage(self._chart_bar(top_df))
            img.width, img.height = 500, 300
            ws.add_image(img, f"A{row+4}")

        for i in range(1, 7):
            ws.column_dimensions[get_column_letter(i)].width = 18

    # ── Sheet 2 ───────────────────────────────────────────────────────────────
    def _build_detailed_analysis(self):
        ws = self.wb.create_sheet("Detailed Analysis")
        ws.sheet_view.showGridLines = False

        summary = self.results.get("summary_by_group", pd.DataFrame())
        if summary.empty:
            ws["A1"] = "No data available."; return

        ws.merge_cells("A1:H1")
        _hdr(ws, 1, 1, f"Performance by {self.group_col.title()}",
             bg=DARK_BLUE, fg=WHITE, size=13, h_align="left")
        ws.row_dimensions[1].height = 30

        headers = list(summary.columns)
        for c, h in enumerate(headers, 1):
            _hdr(ws, 2, c, h, bg=MID_BLUE, fg=WHITE)

        for r_idx, row in enumerate(summary.itertuples(index=False), 3):
            bg = LIGHT_GRAY if r_idx % 2 == 0 else WHITE
            for c_idx, val in enumerate(row, 1):
                col_name = headers[c_idx-1].lower()
                num_fmt, h_align = None, "left"
                if any(k in col_name for k in ["total","average","min","max"]):
                    num_fmt, h_align = "#,##0.00", "right"
                elif "%" in col_name:
                    num_fmt, h_align = '0.00"%"', "right"
                elif "count" in col_name:
                    num_fmt, h_align = "#,##0", "right"
                _val(ws, r_idx, c_idx, val, bg=bg, num_fmt=num_fmt, h_align=h_align)

        total_row = len(summary) + 3
        _hdr(ws, total_row, 1, "TOTAL", bg=DARK_BLUE)
        if "Total" in headers:
            tc = headers.index("Total") + 1
            cell = _val(ws, total_row, tc, round(summary["Total"].sum(), 2),
                        bold=True, bg=DARK_BLUE, num_fmt="#,##0.00", h_align="right")
            cell.font = Font(name="Arial", bold=True, color=WHITE)

        img = XLImage(self._chart_pie(summary))
        img.width, img.height = 420, 340
        ws.add_image(img, f"{get_column_letter(len(headers)+2)}2")
        _auto_width(ws)

    # ── Sheet 3 ───────────────────────────────────────────────────────────────
    def _build_monthly_trend(self):
        ws = self.wb.create_sheet("Monthly Trend")
        ws.sheet_view.showGridLines = False
        trend = self.results["monthly_trend"]

        ws.merge_cells("A1:F1")
        _hdr(ws, 1, 1, "Monthly Revenue Trend",
             bg=DARK_BLUE, fg=WHITE, size=13, h_align="left")
        ws.row_dimensions[1].height = 30

        headers = list(trend.columns)
        for c, h in enumerate(headers, 1):
            _hdr(ws, 2, c, h, bg=MID_BLUE, fg=WHITE)

        for r_idx, row in enumerate(trend.itertuples(index=False), 3):
            bg = LIGHT_GRAY if r_idx % 2 == 0 else WHITE
            for c_idx, val in enumerate(row, 1):
                col_name = headers[c_idx-1].lower()
                num_fmt, h_align = None, "left"
                if "total" in col_name: num_fmt, h_align = "#,##0.00", "right"
                elif "growth" in col_name: num_fmt, h_align = '0.00"%"', "right"
                elif "count" in col_name: num_fmt, h_align = "#,##0", "right"
                _val(ws, r_idx, c_idx, val, bg=bg, num_fmt=num_fmt, h_align=h_align)

        img = XLImage(self._chart_line(trend))
        img.width, img.height = 540, 300
        ws.add_image(img, f"A{len(trend)+5}")
        _auto_width(ws)

    # ── Sheet 4 ───────────────────────────────────────────────────────────────
    def _build_category_breakdown(self):
        ws = self.wb.create_sheet("Category Breakdown")
        ws.sheet_view.showGridLines = False
        cb = self.results["category_breakdown"]

        ws.merge_cells("A1:J1")
        _hdr(ws, 1, 1, "Category Breakdown (Cross-Tab)",
             bg=DARK_BLUE, fg=WHITE, size=13, h_align="left")
        ws.row_dimensions[1].height = 30

        headers = list(cb.columns)
        for c, h in enumerate(headers, 1):
            _hdr(ws, 2, c, str(h), bg=MID_BLUE, fg=WHITE)

        for r_idx, row in enumerate(cb.itertuples(index=False), 3):
            bg = LIGHT_GRAY if r_idx % 2 == 0 else WHITE
            for c_idx, val in enumerate(row, 1):
                num_fmt = "#,##0.00" if c_idx > 1 else None
                _val(ws, r_idx, c_idx, val, bg=bg,
                     num_fmt=num_fmt, h_align="right" if c_idx > 1 else "left")
        _auto_width(ws)

    # ── Sheet 5 ───────────────────────────────────────────────────────────────
    # Raw data row cap — beyond this, nobody scrolls in Excel
    RAW_DATA_ROW_LIMIT = 50_000

    def _build_raw_data(self):
        ws = self.wb.create_sheet("Raw Data")
        ws.sheet_view.showGridLines = True

        df = self.raw_df
        capped = len(df) > self.RAW_DATA_ROW_LIMIT
        if capped:
            df = df.head(self.RAW_DATA_ROW_LIMIT)

        # ── Title banner ──────────────────────────────────────────────────
        title_txt = (
            f"Cleaned Source Data  "
            f"({'first ' + f'{self.RAW_DATA_ROW_LIMIT:,}' + ' of ' if capped else ''}"
            f"{len(self.raw_df):,} rows × {len(self.raw_df.columns)} cols)"
        )
        ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
        _hdr(ws, 1, 1, title_txt, bg=DARK_GRAY, fg=WHITE, size=11, h_align="left")
        ws.row_dimensions[1].height = 26

        # ── Header row (styled) ───────────────────────────────────────────
        for c, col in enumerate(df.columns, 1):
            _hdr(ws, 2, c, col, bg=DARK_GRAY, fg=WHITE, size=10)

        # ── Data rows — bulk append (NO per-cell styling = 30x faster) ───
        # Styling is skipped for data rows to keep large files fast.
        # Header row already provides clear visual separation.
        for row in df.itertuples(index=False):
            ws.append(list(row))

        # ── Freeze header + auto-width ────────────────────────────────────
        ws.freeze_panes = "A3"
        _auto_width(ws, min_w=8, max_w=35)

        if capped:
            logger.info(
                "Raw Data sheet capped at %d rows (dataset has %d rows).",
                self.RAW_DATA_ROW_LIMIT, len(self.raw_df)
            )

    # ── Chart helpers ──────────────────────────────────────────────────────────
    def _chart_bar(self, df: pd.DataFrame) -> io.BytesIO:
        group_col = df.columns[0]
        val_col   = "Total" if "Total" in df.columns else df.columns[1]
        # Drop NaN values before plotting
        df = df[[group_col, val_col]].copy()
        df[val_col] = pd.to_numeric(df[val_col], errors="coerce")
        df = df.dropna(subset=[val_col]).reset_index(drop=True)
        fig, ax = plt.subplots(figsize=(7, 4))
        colors = sns.color_palette("Blues_r", len(df))
        bars = ax.barh(df[group_col], df[val_col], color=colors, edgecolor="white")
        ax.set_xlabel(val_col, fontsize=10)
        ax.set_title(f"Top {len(df)} by {val_col}", fontsize=12, fontweight="bold", pad=12)
        ax.invert_yaxis()
        ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{x:,.0f}"))
        ax.spines[["top","right"]].set_visible(False)
        ax.bar_label(bars, fmt="{:,.0f}", padding=4, fontsize=8)
        fig.tight_layout()
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=130, bbox_inches="tight")
        plt.close(fig); buf.seek(0)
        return buf

    def _chart_pie(self, df: pd.DataFrame) -> io.BytesIO:
        group_col = df.columns[0]
        val_col   = "Total" if "Total" in df.columns else df.columns[1]
        # Drop NaN and zero/negative values — pie chart cannot handle them
        plot_df = df[[group_col, val_col]].copy()
        plot_df[val_col] = pd.to_numeric(plot_df[val_col], errors="coerce")
        plot_df = plot_df.dropna(subset=[val_col])
        plot_df = plot_df[plot_df[val_col] > 0].reset_index(drop=True)
        if plot_df.empty:
            # Return a blank placeholder image if no valid data
            fig, ax = plt.subplots(figsize=(6, 5))
            ax.text(0.5, 0.5, "No data available for chart",
                    ha="center", va="center", fontsize=12, color="gray")
            ax.axis("off")
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=130, bbox_inches="tight")
            plt.close(fig); buf.seek(0)
            return buf
        df = plot_df
        fig, ax = plt.subplots(figsize=(6, 5))
        colors = sns.color_palette("Set2", len(df))
        wedges, _, auto = ax.pie(
            df[val_col], autopct="%1.1f%%", colors=colors,
            startangle=140, pctdistance=0.82,
            wedgeprops=dict(linewidth=0.8, edgecolor="white"),
        )
        for t in auto: t.set_fontsize(8)
        ax.legend(wedges, df[group_col], title=group_col.title(),
                  loc="lower center", bbox_to_anchor=(0.5, -0.18), ncol=3, fontsize=8)
        ax.set_title(f"Share of Total {val_col}", fontsize=11, fontweight="bold", pad=10)
        fig.tight_layout()
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=130, bbox_inches="tight")
        plt.close(fig); buf.seek(0)
        return buf

    def _chart_line(self, df: pd.DataFrame) -> io.BytesIO:
        month_col = df.columns[0]
        val_col   = "Total" if "Total" in df.columns else df.columns[1]
        # Drop NaN values before plotting
        df = df[[month_col, val_col]].copy()
        df[val_col] = pd.to_numeric(df[val_col], errors="coerce")
        df = df.dropna(subset=[val_col]).reset_index(drop=True)
        fig, ax = plt.subplots(figsize=(8, 4))
        ax.plot(df[month_col], df[val_col], marker="o", linewidth=2.5,
                color=f"#{MID_BLUE}", markerfacecolor=f"#{ACCENT}", markersize=7)
        ax.fill_between(df[month_col], df[val_col], alpha=0.12, color=f"#{MID_BLUE}")
        ax.set_title("Monthly Trend", fontsize=12, fontweight="bold", pad=12)
        ax.set_xlabel("Month", fontsize=10)
        ax.set_ylabel(val_col, fontsize=10)
        plt.xticks(rotation=45, ha="right", fontsize=8)
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{x:,.0f}"))
        ax.spines[["top","right"]].set_visible(False)
        ax.grid(axis="y", linestyle="--", alpha=0.4)
        fig.tight_layout()
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=130, bbox_inches="tight")
        plt.close(fig); buf.seek(0)
        return buf
